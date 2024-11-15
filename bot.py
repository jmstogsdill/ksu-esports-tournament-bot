import aiohttp
import aiosqlite # Using this package instead of sqlite for asynchronous processing support
import asyncio
from collections import defaultdict
from datetime import datetime, timedelta
import discord
from discord import AllowedMentions, app_commands
from discord.ext import commands, tasks
from discord.utils import get
from dotenv import load_dotenv, find_dotenv
import itertools
import json
import logging
from openpyxl import load_workbook
import os
import platform
import random
import traceback


load_dotenv(find_dotenv())
intents = discord.Intents.default()
intents.message_content = True
intents.members = True
client = discord.Client(intents=intents)
tree = app_commands.CommandTree(client)


TOKEN = os.getenv('BOT_TOKEN')#Gets the bot's password token from the .env file and sets it to TOKEN.
GUILD = os.getenv('GUILD_TOKEN')#Gets the server's id from the .env file and sets it to GUILD.
# Paths for spreadsheet and SQLite database on the bot host's device
SPREADSHEET_PATH = os.path.abspath(os.getenv('SPREADSHEET_PATH'))
DB_PATH = os.getenv('DB_PATH')
RIOT_API_KEY = os.getenv('RIOT_API_KEY')
TIER_WEIGHT = float(os.getenv('TIER_WEIGHT', 0.7))  # Default value of 0.7 if not specified in .env
ROLE_PREFERENCE_WEIGHT = float(os.getenv('ROLE_PREFERENCE_WEIGHT', 0.3))  # Default value of 0.3 if not specified in .env
TIER_GROUPS = os.getenv('TIER_GROUPS', 'UNRANKED,IRON,BRONZE,SILVER:GOLD,PLATINUM:EMERALD:DIAMOND:MASTER:GRANDMASTER:CHALLENGER') # Setting default tier configuration if left blank in .env


# # Adjust event loop policy for Windows
if platform.system() == "Windows":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    
# Semaphore to limit the number of concurrent API requests to Riot (added to address connection errors pulling player rank with /stats)
api_semaphore = asyncio.Semaphore(5)  # Limit concurrent requests to 5

session = None


# SQLite connection
async def get_db_connection():
    return await aiosqlite.connect(DB_PATH)

# creating db tables if they don't already exist
async def initialize_database():
    async with aiosqlite.connect(DB_PATH) as conn:
        await conn.execute('''
            CREATE TABLE IF NOT EXISTS "PlayerStats" (
                "DiscordID" TEXT NOT NULL UNIQUE,
                "DiscordUsername" TEXT NOT NULL,
                "PlayerRiotID" TEXT UNIQUE,
                "Participation" NUMERIC DEFAULT 0,
                "Wins" INTEGER DEFAULT 0,
                "MVPs" INTEGER DEFAULT 0,
                "ToxicityPoints" NUMERIC DEFAULT 0,
                "GamesPlayed" INTEGER DEFAULT 0,
                "WinRate" REAL,
                "TotalPoints" NUMERIC DEFAULT 0,
                "PlayerTier" INTEGER DEFAULT 0,
                "PlayerRank" TEXT DEFAULT 'UNRANKED',
                "RolePreference" TEXT DEFAULT '55555',
                PRIMARY KEY("DiscordID")
)
        ''')
        await conn.commit()

# On bot ready event
@client.event
async def on_ready():
    global session
    # Initialize aiohttp session when the bot starts (simplified)
    connector = aiohttp.TCPConnector(
        ttl_dns_cache=300,  # Cache DNS resolution for 5 minutes
        ssl=False           # Disable SSL verification (use with caution)
    )
    session = aiohttp.ClientSession(connector=connector)
    
    await initialize_database()
    await tree.sync(guild=discord.Object(GUILD))
    print(f'Logged in as {client.user}')
    
    # Get the guild object
    guild = discord.utils.get(client.guilds, id=int(GUILD))
    if guild is None:
        print(f'Guild with ID {GUILD} not found.')
        return

    # Get bot's role and the roles for Player and Volunteer
    bot_role = discord.utils.get(guild.roles, name=client.user.name)
    player_role = discord.utils.get(guild.roles, name='Player')
    volunteer_role = discord.utils.get(guild.roles, name='Volunteer')

    # Create Player role if it doesn't exist
    if player_role is None:
        player_role = await guild.create_role(name='Player', mentionable=True)
    
    # Create Volunteer role if it doesn't exist
    if volunteer_role is None:
        volunteer_role = await guild.create_role(name='Volunteer', mentionable=True)

    # Adjust Player and Volunteer roles to be below the bot role, if possible
    if bot_role is not None:
        new_position = max(bot_role.position - 1, 1)
        await player_role.edit(position=new_position)
        await volunteer_role.edit(position=new_position)
        
# Safe API call function with retries for better error handling
async def safe_api_call(url, headers):
    max_retries = 3
    retry_delay = 1  # Delay between retries in seconds

    for attempt in range(max_retries):
        try:
            async with api_semaphore:  # Limit concurrent access to the API
                # Using a timeout context inside the aiohttp request
                timeout = aiohttp.ClientTimeout(total=5)  # Set a 5-second total timeout for the request
                async with session.get(url, headers=headers, timeout=timeout) as response:
                    if response.status == 200:
                        return await response.json()
                    elif response.status == 429:  # Rate limit hit
                        retry_after = response.headers.get("Retry-After", 1)
                        print(f"Rate limit reached. Retrying after {retry_after} seconds.")
                        await asyncio.sleep(int(retry_after))  # Wait for specified time
                    else:
                        print(f"Error fetching data: {response.status}, response: {await response.text()}")
                        return None
        except aiohttp.ClientConnectorError as e:
            print(f"Connection error on attempt {attempt + 1}/{max_retries}: {e}")
        except asyncio.TimeoutError as e:
            print(f"Timeout error on attempt {attempt + 1}/{max_retries}: {e}")
        except Exception as e:
            print(f"An unexpected error occurred on attempt {attempt + 1}/{max_retries}: {e}")
            print(traceback.format_exc())  # Print the full traceback for more details

        # Retry if there are remaining attempts
        if attempt < max_retries - 1:
            print(f"Retrying API call in {retry_delay} seconds... (Attempt {attempt + 2}/{max_retries})")
            await asyncio.sleep(retry_delay)

    print("All attempts to connect to the Riot API have failed.")
    return None


"""
update_excel() is a function to update the Excel file / spreadsheet for offline database manipulation.
This implementation (as of 2024-10-25) allows the bot host to update the database by simply altering the
spreadsheet, and vice versa; changes through the bot / db update the spreadsheet at the specified path in .env.

This function will work correctly regardless of whether users rename the template spreadsheet (called "PlayerStats.xlsx"), but do not change the sheet name inside it from "PlayerStats".
In this code, the "workbook" is what we normally call a spreadsheet, and the "sheet name" is just a tab inside that spreadsheet.
So, don't confuse the two and think that you can't change the spreadsheet's file name to something different; you can, as long as you leave the tab/sheet name unaltered and
update the spreadsheet path in ".env".

"""

def update_excel(discord_id, player_data):
    try:
        # Load the workbook and get the correct sheet
        workbook = load_workbook(SPREADSHEET_PATH)
        sheet_name = 'PlayerStats'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f'Sheet {sheet_name} does not exist in the workbook')

        # Check if the player already exists in the sheet
        found = False
        for row in sheet.iter_rows(min_row=2):  # Assuming the first row is headers
            if str(row[0].value) == discord_id:  # Check if Discord ID matches
                # Update only if there's a difference
                for idx, key in enumerate(player_data.keys()):
                    if row[idx].value != player_data[key]:
                        row[idx].value = player_data[key]
                        found = True
                break

        # If player not found, add a new row
        if not found:
            # Find the first truly empty row, ignoring formatting
            empty_row_idx = sheet.max_row + 1
            for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if all(cell.value is None for cell in row):
                    empty_row_idx = i
                    break

            # Insert the new data into the empty row
            for idx, key in enumerate(player_data.keys(), start=1):
                sheet.cell(row=empty_row_idx, column=idx).value = player_data[key]

        # Save the workbook after updates
        workbook.save(SPREADSHEET_PATH)
        print(f"Spreadsheet '{SPREADSHEET_PATH}' has been updated successfully.")

    except Exception as e:
        print(f"Error updating Excel file: {e}")
        


# Updated get_encrypted_summoner_id function
async def get_encrypted_summoner_id(riot_id):
    """
    Fetches the encrypted summoner ID from the Riot API using Riot ID.
    Args:
    - riot_id: The player's Riot ID in 'username#tagline' format.
    Returns:
    - Encrypted summoner ID (summonerId) if successful, otherwise None.
    """
    # Riot ID is expected to be in the format 'username#tagline'
    if '#' not in riot_id:
        return None

    username, tagline = riot_id.split('#', 1)
    url = f"https://americas.api.riotgames.com/riot/account/v1/accounts/by-riot-id/{username}/{tagline}"
    headers = {
        "X-Riot-Token": RIOT_API_KEY
    }

    # Fetch PUUID using the safe API call function
    data = await safe_api_call(url, headers)
    if data:
        puuid = data.get('puuid', None)
        if puuid:
            # Use the PUUID to fetch the encrypted summoner ID
            summoner_url = f"https://na1.api.riotgames.com/lol/summoner/v4/summoners/by-puuid/{puuid}"
            summoner_data = await safe_api_call(summoner_url, headers)
            if summoner_data:
                return summoner_data.get('id', None)  # The summonerId is referred to as `id` in this response

    return None


"""
    Fetches the player's rank from Riot API and updates it in the database.
    Args:
    - conn: The aiosqlite connection object.
    - discord_id: The player's Discord ID.
    - encrypted_summoner_id: The player's encrypted summoner ID.
    - max_retries was added because sometimes the bot failed to connect to the Riot API and properly pull player rank etc (which was leading to rank showing as N/A in /stats.)
      Now, the bot automatically retries the connection several times if this occurs.
"""

async def update_player_rank(conn, discord_id, encrypted_summoner_id):
    await asyncio.sleep(3)  # Initial delay before making the API call

    url = f"https://na1.api.riotgames.com/lol/league/v4/entries/by-summoner/{encrypted_summoner_id}"
    headers = {
        "X-Riot-Token": RIOT_API_KEY
    }

    max_retries = 3

    for attempt in range(max_retries):
        try:
            async with api_semaphore:
                async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=10)) as response:
                    if response.status == 200:
                        data = await response.json()
                        for entry in data:
                            if entry.get("queueType") == "RANKED_SOLO_5x5":
                                rank = entry.get('tier', 'N/A')
                                await conn.execute(
                                    "UPDATE PlayerStats SET PlayerRank = ? WHERE DiscordID = ?",
                                    (rank, discord_id)
                                )
                                await conn.commit()
                                return rank
                        return "UNRANKED"
                    else:
                        print(f"Error fetching player rank: {response.status}, response: {await response.text()}")
        except (aiohttp.ClientConnectionError, aiohttp.ClientResponseError, aiohttp.ClientPayloadError) as e:
            print(f"An error occurred while connecting to the Riot API (attempt {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                print(f"Retrying in 5 seconds (attempt {attempt + 1}/{max_retries})...")
                await asyncio.sleep(5)  # Wait for 5 seconds before retrying

    print("All attempts to connect to the Riot API have failed.")
    return "N/A"  # Return "N/A" if all attempts fail
        


"""
Command to display stats for a given user which simultaneously syncs the user's stats from the database to a spreadsheet (specified in .env) for easy viewing.0
This command pulls and displays some stats from the database, but also makes an API call through update_player_rank() to get the user's updated League of Legends rank.
There is a known error where player rank may display as "N/A" on the command embed due to a connection issue, even if the API key is set up properly. However, typing the command a 1-2 more
times resolves this.
"""
# Command to display stats for a given user
@tree.command(
    name='stats',
    description='Get inhouse stats for a server member who has connected their Riot account with /link.',
    guild=discord.Object(GUILD)
)
async def stats(interaction: discord.Interaction, player: discord.Member):
    # Check if player name is given
    if not player:
        await interaction.response.send_message("Please provide a player name.", ephemeral=True)
        return

    try:
        # Defer the interaction response to prevent timeout
        await interaction.response.defer(ephemeral=True)

        # Update the player's Discord username in the database if needed
        await update_username(player)

        # Connect to the database
        async with aiosqlite.connect(DB_PATH) as conn:
            # Fetch stats from the database
            async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID=?", (str(player.id),)) as cursor:
                player_stats = await cursor.fetchone()

            # If player exists in the database, proceed
            if player_stats:
                riot_id = player_stats[2]  # The Riot ID column from the database

                # Update encrypted summoner ID and rank in the database
                if riot_id:
                    encrypted_summoner_id = await get_encrypted_summoner_id(riot_id)
                    player_rank = await update_player_rank(conn, str(player.id), encrypted_summoner_id) if encrypted_summoner_id else "N/A"
                else:
                    player_rank = "N/A"

                # Create an embed to display player stats
                embed = discord.Embed(
                    title=f"{player.display_name}'s Stats",
                    color=0xffc629  # Hex color #ffc629
                )
                embed.set_thumbnail(url=player.avatar.url if player.avatar else None)

                # Add player stats to the embed
                embed.add_field(name="Riot ID", value=riot_id or "N/A", inline=False)
                embed.add_field(name="Player Rank", value=player_rank, inline=False)
                embed.add_field(name="Participation Points", value=player_stats[3], inline=True)
                embed.add_field(name="Games Played", value=player_stats[7], inline=True)
                embed.add_field(name="Wins", value=player_stats[4], inline=True)
                embed.add_field(name="MVPs", value=player_stats[5], inline=True)
                embed.add_field(name="Win Rate", value=f"{player_stats[8] * 100:.0f}%" if player_stats[8] is not None else "N/A", inline=True)

                # Send the embed as a follow-up response
                await interaction.followup.send(embed=embed, ephemeral=True)

                # Prepare player data dictionary to pass to update_excel
                player_data = {
                    "DiscordID": player_stats[0],
                    "DiscordUsername": player_stats[1],
                    "PlayerRiotID": player_stats[2],
                    "Participation": player_stats[3],
                    "Wins": player_stats[4],
                    "MVPs": player_stats[5],
                    "ToxicityPoints": player_stats[6],
                    "GamesPlayed": player_stats[7],
                    "WinRate": player_stats[8],
                    "PlayerTier": player_stats[10],
                    "PlayerRank": player_rank,
                    "RolePreference": player_stats[12]
                }

                
                # Load the workbook and sheet
                workbook = load_workbook(SPREADSHEET_PATH)
                sheet = workbook.active  # Using the active sheet as the default

                # Check if the player exists in the sheet and if updates are needed
                found = False
                needs_update = False
                for row in sheet.iter_rows(min_row=2):  # Assuming the first row is headers
                    if str(row[0].value) == player_data["DiscordID"]:
                        # Compare each cell to see if an update is needed
                        for key, cell in zip(player_data.keys(), row):
                            if cell.value != player_data[key]:
                                needs_update = True
                                break
                        found = True
                        break

                # If player is not found or an update is needed, update the Excel sheet
                if not found or needs_update:
                    await asyncio.to_thread(update_excel, str(player.id), player_data)
                    await interaction.followup.send(f"Player stats for {player.display_name} have been updated in the Excel sheet.", ephemeral=True)

            else:
                await interaction.followup.send(f"No stats found for {player.display_name}", ephemeral=True)

    except Exception as e:
        # Log the error or handle it appropriately
        print(f"An error occurred: {e}")
        await interaction.followup.send("An unexpected error occurred while fetching player stats.", ephemeral=True)
        

# riot ID linking command, role preference command, and other code added by Jackson 10/18/2024 - may not interact with other code tied to matchmaking with ranks/tiers right now, so can be changed or removed later.
# Dropdown menu for role preference selection

# Riot ID linking command. This is the first command users should type before using other bot features, as it creates a record for them in the database.
@tree.command(
    name='link',
    description="Link your Riot ID to your Discord account.",
    guild=discord.Object(GUILD)
)
async def link(interaction: discord.Interaction, riot_id: str):
    member = interaction.user

    # Riot ID is in the format 'username#tagline', e.g., 'jacstogs#1234'
    if '#' not in riot_id:
        await interaction.response.send_message(
            "Invalid Riot ID format. Please enter your Riot ID in the format 'username#tagline'.",
            ephemeral=True
        )
        return

    # Split the Riot ID into name and tagline
    summoner_name, tagline = riot_id.split('#', 1)
    summoner_name = summoner_name.strip()
    tagline = tagline.strip()

    # Verify that the Riot ID exists using the Riot API
    api_key = os.getenv("RIOT_API_KEY")
    headers = {"X-Riot-Token": api_key}
    url = f"https://americas.api.riotgames.com/riot/account/v1/accounts/by-riot-id/{summoner_name}/{tagline}"

    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, headers=headers) as response:
                if response.status == 200:
                    # Riot ID exists, proceed to link it
                    data = await response.json()  # Get the response data

                    # Debugging: Print the data to see what comes back from the API
                    print(f"Riot API response: {data}")

                    async with aiosqlite.connect(DB_PATH) as conn:
                        # Check if the user already exists in the database
                        async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (str(member.id),)) as cursor:
                            result = await cursor.fetchone()

                        if result:
                            # Update the existing record with the new Riot ID
                            await conn.execute(
                                "UPDATE PlayerStats SET PlayerRiotID = ? WHERE DiscordID = ?",
                                (riot_id, str(member.id))
                            )
                        else:
                            # Insert a new record if the user doesn't exist in the database
                            await conn.execute(
                                "INSERT INTO PlayerStats (DiscordID, DiscordUsername, PlayerRiotID) VALUES (?, ?, ?)",
                                (str(member.id), member.display_name, riot_id)
                            )
                        
                        await conn.commit()

                    await interaction.response.send_message(
                        f"Your Riot ID '{riot_id}' has been successfully linked to your Discord account.",
                        ephemeral=True
                    )
                else:
                    # Riot ID does not exist or other error
                    error_msg = await response.text()
                    print(f"Riot API error response: {error_msg}")
                    await interaction.response.send_message(
                        f"The Riot ID '{riot_id}' could not be found. Please double-check and try again.",
                        ephemeral=True
                    )
    except Exception as e:
        print(f"An error occurred while connecting to the Riot API: {e}")
        await interaction.response.send_message(
            "An unexpected error occurred while trying to link your Riot ID. Please try again later.",
            ephemeral=True
        )


@tree.command(
    name='unlink',
    description="Unlink a player's Riot ID and remove their statistics from the database.",
    guild=discord.Object(GUILD) 
)
@commands.has_permissions(administrator=True)
async def unlink(interaction: discord.Interaction, player: discord.Member):
    try:
        # Store the player to be unlinked for confirmation
        global player_to_unlink
        player_to_unlink = player
        await interaction.response.send_message(f"Type /confirm if you are sure you want to remove {player.display_name}'s record from the bot database.", ephemeral=True)
    
    except commands.MissingPermissions:
        await interaction.response.send_message("You do not have permission to use this command. Only administrators can unlink a player's account.", ephemeral=True)
    
    except Exception as e:
        # Log the error or handle it appropriately
        print(f"An error occurred: {e}")
        await interaction.response.send_message("An unexpected error occurred while unlinking the account.", ephemeral=True)

@tree.command(
    name='confirm',
    description="Confirm the removal of a player's statistics from the database.",
    guild=discord.Object(GUILD)
)
@commands.has_permissions(administrator=True)
async def confirm(interaction: discord.Interaction):
    global player_to_unlink
    try:
        if player_to_unlink:
            # Check if the user exists in the database
            async with aiosqlite.connect(DB_PATH) as conn:
                async with conn.execute("SELECT * FROM PlayerStats WHERE DiscordID = ?", (str(player_to_unlink.id),)) as cursor:
                    player_stats = await cursor.fetchone()

                if player_stats:
                    # Delete user from the database
                    await conn.execute("DELETE FROM PlayerStats WHERE DiscordID = ?", (str(player_to_unlink.id),))
                    await conn.commit()
                    await interaction.response.send_message(f"{player_to_unlink.display_name}'s Riot ID and statistics have been successfully unlinked and removed from the database.", ephemeral=True)
                    player_to_unlink = None
                else:
                    await interaction.response.send_message(f"No statistics found for {player_to_unlink.display_name}. Make sure the account is linked before attempting to unlink.", ephemeral=True)
        else:
            await interaction.response.send_message("No player unlink request found. Please use /unlink first.", ephemeral=True)
    
    except commands.MissingPermissions:
        await interaction.response.send_message("You do not have permission to use this command. Only administrators can confirm the unlinking of a player's account.", ephemeral=True)
    
    except Exception as e:
        # Log the error or handle it appropriately
        print(f"An error occurred: {e}")
        await interaction.response.send_message("An unexpected error occurred while confirming the unlinking of the account.", ephemeral=True)


@tree.command(
    name='resetdb',
    description="Reset player data to defaults, except for ID/rank/role preference information.",
    guild=discord.Object(GUILD)
)
async def resetdb(interaction: discord.Interaction):
    # Only the server owner can use this command
    if interaction.user != interaction.guild.owner:
        await interaction.response.send_message(
            "You do not have permission to use this command. Only the server owner can reset the database.",
            ephemeral=True
        )
        return

    # Send confirmation message to the server owner
    await interaction.response.send_message(
        "You are about to reset the player database to default values for participation, wins, MVPs, toxicity points, games played, win rate, and total points (excluding rank, tier, and role preferences). "
        "Please type /resetdb again within the next 10 seconds to confirm.",
        ephemeral=True
    )

    def check(res: discord.Interaction):
        # Check if the command is resetdb and if it's the same user who issued the original command
        return res.command.name == 'resetdb' and res.user == interaction.user

    try:
        # Wait for the confirmation within 10 seconds
        response = await client.wait_for('interaction', timeout=10.0, check=check)

        # If the confirmation is received, proceed with resetting the database
        async with aiosqlite.connect(DB_PATH) as conn:
            await conn.execute("""
                UPDATE PlayerStats
                SET
                    Participation = 0,
                    Wins = 0,
                    MVPs = 0,
                    ToxicityPoints = 0,
                    GamesPlayed = 0,
                    WinRate = NULL,
                    TotalPoints = 0
            """)
            await conn.commit()

        # Send a follow-up message indicating the reset was successful
        await response.followup.send(
            "The player database has been successfully reset to default values, excluding rank, tier, and role preferences.",
            ephemeral=True
        )

    except asyncio.TimeoutError:
        # If no confirmation is received within 10 seconds, send a follow-up message indicating timeout
        await interaction.followup.send(
            "Reset confirmation timed out. Please type /resetdb again if you still wish to reset the database.",
            ephemeral=True
        )

        

# Role preference command
class RolePreferenceView(discord.ui.View):
    def __init__(self, member_id, initial_values):
        super().__init__(timeout=60)
        self.member_id = member_id
        self.values = initial_values  # Use the initial preferences from the database
        self.embed_message = None  # Track the embed message to edit later
        roles = ["Top", "Jungle", "Mid", "Bot", "Support"]
        for role in roles:
            self.add_item(RolePreferenceDropdown(role, self))  # Pass the view to the dropdown

    async def update_embed_message(self, interaction):
        # Create an embed to display role preferences
        embed = discord.Embed(title="Role Preferences", color=0xffc629)
        for role, value in self.values.items():
            embed.add_field(name=role, value=f"Preference: {value}", inline=False)

        # Send or edit the ephemeral embed message with the updated preferences
        if self.embed_message is None:
            self.embed_message = await interaction.followup.send(embed=embed, ephemeral=True)
        else:
            await self.embed_message.edit(embed=embed)


class RolePreferenceDropdown(discord.ui.Select):
    def __init__(self, role, parent_view: RolePreferenceView):
        self.role = role
        self.parent_view = parent_view  # Reference to the parent view
        options = [
            discord.SelectOption(label=str(i), value=str(i)) for i in range(1, 6)
        ]
        super().__init__(
            placeholder=f"Select your matchmaking priority for {role}",
            min_values=1,
            max_values=1,
            options=options,
        )

    async def callback(self, interaction: discord.Interaction):
        # Update the selected value in the parent view's `values` dictionary
        self.parent_view.values[self.role] = int(self.values[0])

        # Concatenate the role preferences into a single string
        role_pref_string = ''.join(str(self.parent_view.values[role]) for role in ["Top", "Jungle", "Mid", "Bot", "Support"])

        # Update the database with the new role preferences
        async with aiosqlite.connect(DB_PATH) as conn:
            await conn.execute(
                "UPDATE PlayerStats SET RolePreference = ? WHERE DiscordID = ?",
                (role_pref_string, str(self.parent_view.member_id))
            )
            await conn.commit()

        # Acknowledge interaction
        await interaction.response.defer()  # Acknowledge the interaction without updating the message

        # Update the role preferences embed
        await self.parent_view.update_embed_message(interaction)


@tree.command(
    name='rolepreference',
    description="Set your role preferences for matchmaking.",
    guild=discord.Object(GUILD)
)
async def rolepreference(interaction: discord.Interaction):
    member = interaction.user

    # Check if the user has the Player or Volunteer role
    if not any(role.name in ["Player", "Volunteer"] for role in member.roles):
        await interaction.response.send_message("You must have the Player or Volunteer role to set role preferences.", ephemeral=True)
        return

    # Check if the user is in the database and retrieve their current preferences
    async with aiosqlite.connect(DB_PATH) as conn:
        async with conn.execute("SELECT RolePreference FROM PlayerStats WHERE DiscordID = ?", (str(member.id),)) as cursor:
            user_data = await cursor.fetchone()

        if not user_data:
            await interaction.response.send_message("You need to link your Riot ID using /link before setting role preferences.", ephemeral=True)
            return

    # Convert the existing role preferences into a dictionary
    role_pref_string = user_data[0]
    initial_values = {
        "Top": int(role_pref_string[0]),
        "Jungle": int(role_pref_string[1]),
        "Mid": int(role_pref_string[2]),
        "Bot": int(role_pref_string[3]),
        "Support": int(role_pref_string[4])
    }

    # Create the view with initial values and send initial response
    view = RolePreferenceView(member.id, initial_values)
    await interaction.response.send_message(
        "Please select your roles in order of preference, with 1 being the most preferred:", 
        view=view,
        ephemeral=True
    )



# code to calculate and update winrate in database
async def update_win_rate(discord_id):
    async with await get_db_connection() as conn:
        async with conn.execute("SELECT Wins, GamesPlayed FROM PlayerStats WHERE DiscordID = ?", (discord_id,)) as cursor:
            result = await cursor.fetchone()
    if result:
        wins, games_played = result
        win_rate = (wins / games_played) * 100 if games_played > 0 else 0
        await conn.execute("UPDATE PlayerStats SET WinRate = ? WHERE DiscordID = ?", (win_rate, discord_id))
        await conn.commit()




# end of aforementioned code added by Jackson


#Player class.
class Player:
    def __init__(self, tier, username, discord_id, top_priority, jungle_priority, mid_priority, bot_priority, support_priority):
        self.tier = tier
        self.username = username
        self.discord_id = discord_id
        self.top_priority = top_priority
        self.jungle_priority = jungle_priority
        self.mid_priority = mid_priority
        self.bot_priority = bot_priority
        self.support_priority = support_priority
        self.random_factor = random.uniform(0.9, 1.1)  # Random factor to slightly vary the player's weight for matchmaking
        
    def calculate_weight(self, tier_weight, role_preference_weight):
        # Calculate weight based on tier, role preference, and a random factor
        base_weight = self.tier * tier_weight + (
            self.top_priority +
            self.jungle_priority +
            self.mid_priority +
            self.bot_priority +
            self.support_priority
        ) * role_preference_weight
        
        return base_weight * self.random_factor  # Add randomness to the final weight
        

#Team class.
class Team:
    def __init__(self, top_laner, jungle, mid_laner, bot_laner, support):
        self.top_laner = top_laner
        self.jungle = jungle
        self.mid_laner = mid_laner
        self.bot_laner = bot_laner
        self.support = support

    def __str__(self):
        return f"Top Laner: {self.top_laner.username} priority: {self.top_laner.top_priority} (Tier {self.top_laner.tier})\nJungle: {self.jungle.username} priority:{self.jungle.jungle_priority} \
              (Tier {self.jungle.tier})\nMid Laner: {self.mid_laner.username} priority: {self.mid_laner.mid_priority} (Tier {self.mid_laner.tier})\nBot Laner: {self.bot_laner.username} \
                  priority: {self.bot_laner.bot_priority} (Tier {self.bot_laner.tier})\nSupport: {self.support.username} priority: {self.support.support_priority} (Tier {self.support.tier})"


#Check-in button class for checking in to tournaments.
class CheckinButtons(discord.ui.View):
    # timeout after 900 seconds = end of 15-minute check-in period
    def __init__(self, *, timeout = 900):
        super().__init__(timeout = timeout)
    """
    This button is a green button that is called check in
    When this button is pulled up, it will show the text "Check-In"

    The following output when clicking the button is to be expected:
    If the user already has the player role, it means that they are already checked in.
    If the user doesn't have the player role, it will give them the player role. 
    """
    @discord.ui.button(
            label = "Check-In",
            style = discord.ButtonStyle.green)
    async def checkin(self, interaction: discord.Interaction, button: discord.ui.Button):

        player = get(interaction.guild.roles, name = 'Player')
        member = interaction.user

        if player in member.roles:
            await interaction.response.edit_message(view = self)
            await interaction.followup.send('You have already checked in.', ephemeral=True)
            return "Is already checked in"
        await member.add_roles(player)
        await interaction.response.edit_message(view = self)
        await interaction.followup.send('You have checked in!', ephemeral = True)
        return "Checked in"        

    """
    This button is the leave button. It is used for if the player checked in but has to leave
    The following output is to be expected:

    If the user has the player role, it will remove it and tell the player that it has been removed
    If the user does not have the player role, it will tell them to check in first.
    """
    @discord.ui.button(
            label = "Leave",
            style = discord.ButtonStyle.red)
    async def leave(self, interaction: discord.Interaction, button: discord.ui.Button):

        player = get(interaction.guild.roles, name = 'Player')
        member = interaction.user

        if player in member.roles:
            await member.remove_roles(player)
            await interaction.response.edit_message(view = self)
            await interaction.followup.send('Sorry to see you go.', ephemeral = True)
            return "Role Removed"
        await interaction.response.edit_message(view = self)
        await interaction.followup.send('You have not checked in. Please checkin first', ephemeral = True)
        return "Did not check in yet"

class volunteerButtons(discord.ui.View):
    # timeout after 900 seconds = end of 15-minute volunteer period
    def __init__(self, *, timeout = 900):
        super().__init__(timeout = timeout)
    """
    This button is a green button that is called check in
    When this button is pulled up, it will show the text "Volunteer"

    The following output when clicking the button is to be expected:
    If the user already has the volunteer role, it means that they are already volunteered.
    If the user doesn't have the volunteer role, it will give them the volunteer role. 
    """
    @discord.ui.button(
            label = "Volunteer",
            style = discord.ButtonStyle.green)
    async def checkin(self, interaction: discord.Interaction, button: discord.ui.Button):

        player = get(interaction.guild.roles, name = 'Player')
        volunteer = get(interaction.guild.roles, name = 'Volunteer')
        member = interaction.user

        if player in member.roles:
            await member.remove_roles(player)
        if volunteer in member.roles:
            await interaction.response.edit_message(view = self)
            await interaction.followup.send('You have already volunteered to sit out, if you wish to rejoin click rejoin.', ephemeral=True)
            return "Is already checked in"
        await member.add_roles(volunteer)
        await interaction.response.edit_message(view = self)
        await interaction.followup.send('You have volunteered to sit out!', ephemeral = True)
        return "Checked in"        

    """
    This button is the leave button. It is used for if the player who has volunteer wants to rejoin
    The following output is to be expected:

    If the user has the player role, it will remove it and tell the player that it has been removed
    If the user does not have the volunteer role, it will tell them to volunteer first.
    """
    @discord.ui.button(
            label = "Rejoin",
            style = discord.ButtonStyle.red)
    async def leave(self, interaction: discord.Interaction, button: discord.ui.Button):
        
        player = get(interaction.guild.roles, name = 'Player')
        volunteer = get(interaction.guild.roles, name = 'Volunteer')
        member = interaction.user

        if volunteer in member.roles:
            await member.remove_roles(volunteer)
            await member.add_roles(player)
            await interaction.response.edit_message(view = self)
            await interaction.followup.send('Welcome back in!', ephemeral = True)
            return "Role Removed"
        await interaction.response.edit_message(view = self)
        await interaction.followup.send('You have not volunteered to sit out, please volunteer to sit out first.', ephemeral = True)
        return "Did not check in yet"

    
# Function to update Discord username in the database if it's been changed
async def update_username(player: discord.Member):
    try:
        async with aiosqlite.connect(DB_PATH) as conn:
            # Fetch the player's current data from the database
            async with conn.execute("SELECT DiscordUsername FROM PlayerStats WHERE DiscordID=?", (str(player.id),)) as cursor:
                player_stats = await cursor.fetchone()

            # If player exists in the database and the username is outdated, update it
            if player_stats:
                stored_username = player_stats[0]
                current_username = player.display_name

                if stored_username != current_username:
                    await conn.execute(
                        "UPDATE PlayerStats SET DiscordUsername = ? WHERE DiscordID = ?",
                        (current_username, str(player.id))
                    )
                    await conn.commit()
                    print(f"Updated username in database for {player.id} from '{stored_username}' to '{current_username}'.")

    except Exception as e:
        # Log the error or handle it appropriately
        print(f"An error occurred while updating username: {e}")


async def update_points(members):
    async with aiosqlite.connect(DB_PATH) as conn:
        not_found_users = []
        updated_users = []

        # Iterate through all members with Player or Volunteer roles
        for member in members:
            async with conn.execute("SELECT Participation, GamesPlayed FROM PlayerStats WHERE DiscordID = ?", (str(member.id),)) as cursor:
                result = await cursor.fetchone()

            if result:
                participation, games_played = result

                # Check if the member has the Player or Volunteer role
                if any(role.name == "Player" for role in member.roles):
                    # Update both Participation and GamesPlayed for Players
                    await conn.execute(
                        "UPDATE PlayerStats SET Participation = ?, GamesPlayed = ? WHERE DiscordID = ?",
                        (participation + 1, games_played + 1, str(member.id))
                    )
                    updated_users.append(member.display_name)
                elif any(role.name == "Volunteer" for role in member.roles):
                    # Update only Participation for Volunteers
                    await conn.execute(
                        "UPDATE PlayerStats SET Participation = ? WHERE DiscordID = ?",
                        (participation + 1, str(member.id))
                    )
                    updated_users.append(member.display_name)
            else:
                # Add users who are not found in the database to the list
                not_found_users.append(member.display_name)

        await conn.commit()

    return {"success": updated_users, "not_found": not_found_users}
    
async def update_toxicity(member):
    async with aiosqlite.connect(DB_PATH) as conn:
        # Attempt to find the user in the PlayerStats table
        async with conn.execute("SELECT ToxicityPoints FROM PlayerStats WHERE DiscordID = ?", (str(member.id),)) as cursor:
            result = await cursor.fetchone()

        if result:
            toxicity_points = result[0]
            # Increment the ToxicityPoints and update TotalPoints accordingly
            await conn.execute(
                """
                UPDATE PlayerStats 
                SET ToxicityPoints = ?, TotalPoints = (Participation + Wins - ?)
                WHERE DiscordID = ?
                """,
                (toxicity_points + 1, toxicity_points + 1, str(member.id))
            )
            await conn.commit()
            return True  # Successfully updated user

        return False  # User not found

async def check_winners_in_db(winners):
    not_found_users = []
    async with aiosqlite.connect(DB_PATH) as conn:
        for winner in winners:
            # Check if the player exists in the database
            async with conn.execute("SELECT Wins, GamesPlayed FROM PlayerStats WHERE DiscordID = ?", (str(winner.id),)) as cursor:
                result = await cursor.fetchone()

            if not result:
                # Add to not found list
                not_found_users.append(winner)
    
async def update_wins(winners):
    async with aiosqlite.connect(DB_PATH) as conn:
        for winner in winners:
            # Since we already checked for existence, we can directly update
            async with conn.execute("SELECT Wins, GamesPlayed FROM PlayerStats WHERE DiscordID = ?", (str(winner.id),)) as cursor:
                result = await cursor.fetchone()
                if result:
                    wins, games_played = result
                    # Update the Wins and GamesPlayed for the player
                    await conn.execute(
                        "UPDATE PlayerStats SET Wins = ?, GamesPlayed = ? WHERE DiscordID = ?",
                        (wins + 1, games_played + 1, str(winner.id))
                    )

        await conn.commit()

#Command to start check-in
@tree.command(
    name = 'checkin',
    description = 'Initiate tournament check-in',
    guild = discord.Object(GUILD))
async def checkin(interaction: discord.Interaction):
    player = interaction.user
    
    # Upon checking in, update the player's Discord username in the database if it has been changed
    await update_username(player)
    
    view = CheckinButtons()
    await interaction.response.send_message('Check-in for the tournament has started! You have 15 minutes to check in.', view = view)

#Command to start check for volunteers to sit out of a match.
@tree.command(
    name = 'sitout',
    description = 'Initiate check for volunteers to sit out from a match',
    guild = discord.Object(GUILD))
async def sitout(interaction: discord.Interaction):
    player = interaction.user
    
    # Upon volunteering to sit out for a round, update the player's Discord username in the database if it has been changed.
    # This may be redundant if users are eventually restricted from volunteering before they've checked in.
    await update_username(player)
    
    view = volunteerButtons()
    await interaction.response.send_message('The Volunteer check has started! You have 15 minutes to volunteer if you wish to sit out.', view = view)

@tree.command(
    name='toxicity',
    description='Give a user a point of toxicity (subtracting 1 from their total League inhouse points).',
    guild=discord.Object(GUILD)
)
@commands.has_permissions(administrator=True)
async def toxicity(interaction: discord.Interaction, member: discord.Member):
    try:
        # Defer the response to prevent interaction timeout
        await interaction.response.defer(ephemeral=True)

        # Update the toxicity points in the database
        found_user = await update_toxicity(member)

        if found_user:
            await interaction.followup.send(f"{member.display_name}'s toxicity points have been updated.")
        else:
            await interaction.followup.send(f"{member.display_name} could not be found in the database.")
    except commands.MissingPermissions:
        await interaction.response.send_message("You do not have permission to use this command. Only administrators can give toxicity points.", ephemeral=True)
    except Exception as e:
        print(f'An error occurred: {e}')
        await interaction.followup.send("An unexpected error occurred while updating toxicity points.", ephemeral=True)

@tree.command(
    name='win',
    description="Update the number of wins for each player on the winning team.",
    guild=discord.Object(GUILD)
)
@commands.has_permissions(administrator=True)
async def win(interaction: discord.Interaction, match_number: str, lobby_number: str, team: str):
    try:
        # Check if the team is either "red" or "blue"
        if team.lower() not in ['red', 'blue']:
            await interaction.response.send_message(
                "Error: Please specify a valid team name (`red` or `blue`).",
                ephemeral=True
            )
            return

        # Ensure that the match and lobby exist by checking against the matchmaking result
        match_key = f"match_{match_number}_lobby_{lobby_number}"
        if match_key not in active_matches:
            await interaction.response.send_message(
                "Error: No match found. Please run `/matchmake` for this match and lobby first.",
                ephemeral=True
            )
            return

        # Get the winning team's players
        winning_team = active_matches[match_key][team.lower()]

        # Update wins for each player on the winning team
        await interaction.response.defer(ephemeral=True)

        not_found_users = await check_winners_in_db(winning_team)

        if not_found_users:
            missing_users = ", ".join([user.username for user in not_found_users])
            await interaction.followup.send(
                f"The following players could not be found in the database, so no wins were updated: \n{missing_users}"
            )
        else:
            await update_wins(winning_team)
            await interaction.followup.send("All winners' 'win' points have been updated.")

    except commands.MissingPermissions:
        await interaction.response.send_message(
            "You do not have permission to use this command. Only administrators can update players' wins.",
            ephemeral=True
        )

    except Exception as e:
        print(f'An error occurred: {e}')
        await interaction.followup.send(
            "An error occurred while updating wins.",
            ephemeral=True
        )


       
# Slash command to remove all users from the Player and Volunteer roles.
@tree.command(
    name='clear',
    description='Remove all users from Player and Volunteer roles.',
    guild=discord.Object(GUILD)
)
async def remove(interaction: discord.Interaction):
    try:
        player = get(interaction.guild.roles, name='Player')
        volunteer = get(interaction.guild.roles, name='Volunteer')

        # Acknowledge the interaction to avoid timeouts.
        await interaction.response.defer(ephemeral=True)

        permission_issue = False

        # Iterate through all members and attempt to remove roles
        for user in interaction.guild.members:
            try:
                if player in user.roles:
                    await user.remove_roles(player)
                if volunteer in user.roles:
                    await user.remove_roles(volunteer)
            except discord.Forbidden:
                # If a permission error occurs, set the flag
                permission_issue = True
                print(f"Could not remove roles from {user.display_name}. Check role hierarchy.")

        # Prepare the response message
        if permission_issue:
            response_message = (
                "Attempted to remove Player and Volunteer roles from all users.\n"
                "Some users could not be updated due to role hierarchy or missing permissions."
            )
        else:
            response_message = "All users have been successfully removed from Player and Volunteer roles."

        # Send the follow-up message to the user
        await interaction.followup.send(response_message, ephemeral=True)

    except Exception as e:
        print(f'An error occurred: {e}')
        await interaction.followup.send("An unexpected error occurred while removing roles from users.", ephemeral=True)

#Slash command to find and count all of the players and volunteers
@tree.command(
        name='players',
        description='Find all players and volunteers currently enrolled in the game',
        guild = discord.Object(GUILD))
async def players(interaction: discord.Interaction):
    try:
        player_users = []
        player = get(interaction.guild.roles, name = 'Player')
        for user in interaction.guild.members:
            if player in user.roles:
                player_users.append(user.name)
        
        #Finds all volunteers in discord, adds them to a list
        volunteer_users = []
        volunteer = get(interaction.guild.roles, name = 'Volunteer')
        for user in interaction.guild.members:
            if volunteer in user.roles:
                volunteer_users.append(user.name)

        player_count = sum(1 for user in interaction.guild.members if player in user.roles)
        volunteer_count = sum(1 for user in interaction.guild.members if volunteer in user.roles)

        #Embed to display all users who volunteered to sit out.
        embedPlayers = discord.Embed(color = discord.Color.green(), title = 'Total Players')
        embedPlayers.set_footer(text = f'Total players: {player_count}')
        for pl in player_users:
            embedPlayers.add_field(name = '', value = pl)

        embedVolunteers = discord.Embed(color = discord.Color.orange(), title = 'Total Volunteers')
        embedVolunteers.set_footer(text = f'Total volunteers: {volunteer_count}')
        for vol in volunteer_users:
            embedVolunteers.add_field(name = '', value = vol)
        
        next_increment = 10 - (player_count % 10)
        message = ''
        if player_count == 10:
            message += "There is a full lobby with 10 players!"
            embedMessage = discord.Embed(color = discord.Color.dark_green(), title = 'Players/Volunteers')
            embedMessage.add_field(name = '', value = message)
        elif next_increment==10 and player_count!=0:
            message += "There are multiple full lobbies ready!"
            embedMessage = discord.Embed(color = discord.Color.dark_green(), title = 'Players/Volunteers')
            embedMessage.add_field(name = '', value = message)
        elif player_count < 10:
            message += "A full lobby requires at least 10 players!"
            embedMessage = discord.Embed(color = discord.Color.dark_red(), title = 'Players/Volunteers')
            embedMessage.add_field(name = '', value = message)
        else:
            message += f"For a full lobby {next_increment} players are needed or {10-next_increment} volunteers are needed!"
            embedMessage = discord.Embed(color = discord.Color.dark_red(), title = 'Players/Volunteers')
            embedMessage.add_field(name = '', value = message)

        await interaction.response.send_message(embeds = [embedMessage, embedPlayers, embedVolunteers])
    except Exception as e:
        print(f'An error occured: {e}')

@tree.command(
    name='points',
    description='Give participation points to all players and volunteers.',
    guild=discord.Object(GUILD)
)
@commands.has_permissions(administrator=True)
async def points(interaction: discord.Interaction):
    try:
        # Finds all players and volunteers in Discord, adds them to respective lists
        player_role = get(interaction.guild.roles, name='Player')
        volunteer_role = get(interaction.guild.roles, name='Volunteer')

        players_and_volunteers = [
            member for member in interaction.guild.members
            if player_role in member.roles or volunteer_role in member.roles
        ]

        await interaction.response.defer(ephemeral=True)

        # Update points for all players and volunteers
        update_result = await update_points(players_and_volunteers)

        # Send followup messages
        if update_result["success"]:
            updated_users = ', '.join(update_result["success"])
            await interaction.followup.send(f"Participation points have been updated for the following users: {updated_users}.", ephemeral=True)
            print(f"Participation points have been updated for the following users: {updated_users}")
        
        if update_result["not_found"]:
            not_found_users = ', '.join(update_result["not_found"])
            await interaction.followup.send(f"The following users were not found in the database: {not_found_users}.", ephemeral=True)
            print(f"The following users were not found in the database: {not_found_users}")
            
    except commands.MissingPermissions:
        await interaction.response.send_message("You do not have permission to use this command. Only administrators can update players' participation points.", ephemeral=True)
        
    except Exception as e:
        print(f'An error occurred: {e}')
        await interaction.followup.send("An unexpected error occurred while updating participation points.", ephemeral=True)

# Parse TIER_GROUPS from the .env file
def parse_tier_groups(tier_groups_string):
    tier_mapping = {}
    groups = tier_groups_string.split(':')
    for tier_index, group in enumerate(groups, start=1):
        ranks = group.split(',')
        for rank in ranks:
            tier_mapping[rank.strip().upper()] = tier_index
    return tier_mapping

TIER_MAPPING = parse_tier_groups(TIER_GROUPS)
active_matches = {}  # Global dictionary to store match and lobby data after `/matchmake`

@tree.command(
    name='matchmake',
    description="Form teams for all players enrolled in the game",
    guild=discord.Object(GUILD)
)
async def matchmake(interaction: discord.Interaction, match_number: str, lobby_number: str):
    try:
        # Find all players in Discord who have the "Player" role
        player_users = []
        player_role = get(interaction.guild.roles, name='Player')
        for user in interaction.guild.members:
            if player_role in user.roles:
                player_users.append(user)

        # Find all volunteers in Discord who have the "Volunteer" role
        volunteer_users = []
        volunteer_role = get(interaction.guild.roles, name='Volunteer')
        for user in interaction.guild.members:
            if volunteer_role in user.roles:
                volunteer_users.append(user)

        # Check if the number of players is valid for matchmaking
        if len(player_users) % 10 != 0:
            await interaction.response.send_message(
                "Error: The number of players must be a multiple of 10.", ephemeral=True)
            return

        # Retrieve player stats from the database
        matched_players = []
        async with aiosqlite.connect(DB_PATH) as conn:
            for player in player_users:
                async with conn.execute("SELECT DiscordID, PlayerRiotID, PlayerRank, PlayerTier, RolePreference FROM PlayerStats WHERE DiscordID = ?", (str(player.id),)) as cursor:
                    player_data = await cursor.fetchone()
                    if player_data:
                        discord_id, riot_id, player_rank, player_tier, role_pref = player_data
                        # Assign tier dynamically based on rank using TIER_MAPPING
                        player_tier = TIER_MAPPING.get(player_rank.upper(), len(TIER_MAPPING) + 1)
                        matched_players.append(Player(
                            tier=player_tier,
                            username=player.display_name,
                            discord_id=discord_id,
                            role_preference=role_pref
                        ))

        if len(matched_players) != len(player_users):
            await interaction.response.send_message(
                "Error: Some players are missing from the database.", ephemeral=True)
            return

        # Create the best teams based on matchmaking criteria
        best_teams = await create_best_teams(matched_players)

        if not best_teams:
            await interaction.response.send_message(
                "Error: Unable to create balanced teams.", ephemeral=True)
            return

        # Prepare embeds for teams and volunteers
        embed_lobby = discord.Embed(color=discord.Color.from_rgb(255, 198, 41), title=f'Lobby {lobby_number} - Match: {match_number}')
        embed_lobby.add_field(name='Roles', value='Top\nJungle\nMid\nBot\nSupport', inline=True)
        embed_lobby.add_field(name='Red Team', value='\n'.join(player.username for player in best_teams[0].red_team), inline=True)
        embed_lobby.add_field(name='Blue Team', value='\n'.join(player.username for player in best_teams[0].blue_team), inline=True)

        embed_volunteers = discord.Embed(color=discord.Color.blurple(), title=f'Volunteers - Match: {match_number}')
        if volunteer_users:
            embed_volunteers.add_field(name='', value='\n'.join(vol.display_name for vol in volunteer_users))
        else:
            embed_volunteers.add_field(name='', value='No volunteers.')

        # Send the embeds
        await interaction.response.send_message(embeds=[embed_lobby, embed_volunteers])

        # Store the match and lobby information for later reference in `/win`
        match_key = f"match_{match_number}_lobby_{lobby_number}"
        active_matches[match_key] = {
            'red': best_teams[0].red_team,
            'blue': best_teams[0].blue_team
        }

    except Exception as e:
        print(f'An error occurred: {e}')
        await interaction.response.send_message(
            "An unexpected error occurred while forming teams.", ephemeral=True
        )

async def create_best_teams(players):
    if len(players) % 10 != 0:
        return None

    sorted_players = sorted(players, key=lambda x: x.tier)
    group_size = 10
    num_groups = (len(sorted_players) + group_size - 1) // group_size
    best_teams = []

    for i in range(num_groups):
        group = sorted_players[i * group_size: (i + 1) * group_size]
        if len(group) == group_size:
            teams = await create_best_teams_helper(group)
            if teams:
                best_teams.extend(teams)

    return best_teams

async def create_randomized_teams(players):
    # Randomly shuffle players to introduce randomness in the matchmaking
    random.shuffle(players)
    
    # Split players into groups of 10 to form teams
    group_size = 10
    num_groups = len(players) // group_size
    teams = []

    for i in range(num_groups):
        group = players[i * group_size: (i + 1) * group_size]
        
        # Generate balanced teams with added randomness
        best_team = await create_best_teams_helper(group)
        if best_team:
            teams.append(best_team)

    return teams



async def create_best_teams_helper(players):
    if len(players) != 10:
        return None

    lowest_score = float('inf')
    best_team = None

    # Randomize permutations to create different results for the same set of players
    permutations = list(itertools.permutations(players, 5))
    random.shuffle(permutations)

    for team1_players in permutations:
        team1 = list(team1_players)
        team2 = [player for player in players if player not in team1]

        # Calculate score difference with some randomness
        score_diff = await calculate_score_diff(Team(*team1), Team(*team2), tier_weight=TIER_WEIGHT)

        # Calculate team priorities with randomness
        team1_priority = await calculate_team_priority(*team1, role_preference_weight=ROLE_PREFERENCE_WEIGHT)
        team2_priority = await calculate_team_priority(*team2, role_preference_weight=ROLE_PREFERENCE_WEIGHT)

        total_score = score_diff * TIER_WEIGHT + (team1_priority + team2_priority) * ROLE_PREFERENCE_WEIGHT

        if total_score < lowest_score:
            lowest_score = total_score
            best_team = Team(team1, team2)

    return best_team



async def calculate_team_priority(top, jungle, mid, bot, support):
    # Calculating the total priority score for the team by adding the squares of each player's role priority.
    # The lower the number, the closer the players are to their most preferred roles.
    priority = 0
    priority += top.top_priority ** 2
    priority += jungle.jungle_priority ** 2
    priority += mid.mid_priority ** 2
    priority += bot.bot_priority ** 2
    priority += support.support_priority ** 2
    return priority


#calculates the score difference in tier for any 2 given teams (i.e. how well it's balanced with regard to tier)
async def calculate_score_diff(team1, team2):
    # Calculating the tier score difference for each role between two teams.
    diff = (team1.top_laner.tier - team2.top_laner.tier) ** 2 + \
           (team1.mid_laner.tier - team2.mid_laner.tier) ** 2 + \
           (team1.bot_laner.tier - team2.bot_laner.tier) ** 2 + \
           (team1.jungle.tier - team2.jungle.tier) ** 2 + \
           (team1.support.tier - team2.support.tier) ** 2
    return diff


def has_roles(team, required_roles):
    # Create a set of roles present in the team
    team_roles = {player.role for player in team}
    # Check if all required roles are in the team_roles set
    return all(role in team_roles for role in required_roles)



# Group players by tier (or nearby tiers)
def group_players_by_rank(players):
    tier_groups = {}
    
    # Group players into their respective tiers
    for player in players:
        if player.tier not in tier_groups:
            tier_groups[player.tier] = []
        tier_groups[player.tier].append(player)
    
    return tier_groups



# code for MVP voting command/functions added by Jackson -- updated morning of 10/18/2024
# Dict that tracks votes after someone initiates an MVP voting session with /votemvp
votes = defaultdict(int)
has_voted = set()  # Tracks players who've already voted
voting_in_progress = False  # Checks if an MVP voting session is already running
mvp_updates_today = 0  # Track the number of MVP updates today (max 3 per day)

# Start an MVP voting period
@tasks.loop(count=1)
async def start_voting(initial_message: discord.Message):
    global voting_in_progress, mvp_updates_today
    voting_in_progress = True

    # Send countdown warnings
    await asyncio.sleep(120)  # Wait for 2 minutes
    await initial_message.channel.send("There are only **3 minutes** remaining for MVP voting!", ephemeral=False)
    
    await asyncio.sleep(60)  # Wait for 1 minute
    await initial_message.channel.send("There are only **2 minutes** remaining for MVP voting!", ephemeral=False)
    
    await asyncio.sleep(60)  # Wait for 1 minute
    await initial_message.channel.send("There is only **1 minute** remaining for MVP voting!", ephemeral=False)

    await asyncio.sleep(60)  # Wait for the final minute to end

    async with await get_db_connection() as conn:
        # Calculate MVP based on votes
        if votes:
            max_votes = max(votes.values())
            mvp_candidates = [player for player, count in votes.items() if count == max_votes]
            
            # Update MVPs in the database
            for mvp in mvp_candidates:
                await conn.execute("UPDATE PlayerStats SET MVPs = MVPs + 1 WHERE DiscordUsername = ?", (mvp,))
            await conn.commit()

            # Prepare MVP result message
            if len(mvp_candidates) == 1:
                await initial_message.channel.send(f"🎉 {mvp_candidates[0]} has been voted the MVP of this round! 🎉", ephemeral=False)
            else:
                mvp_list = ", ".join(mvp_candidates)
                await initial_message.channel.send(f"🎉 The MVP(s) with the highest votes are: {mvp_list}. 🎉", ephemeral=False)

            mvp_updates_today += 1
        else:
            await initial_message.channel.send("No votes were cast. No MVP this round.", ephemeral=False)
    
    votes.clear()
    has_voted.clear()
    voting_in_progress = False

# /votemvp command
@tree.command(
    name='votemvp',
    description='Vote for the MVP of your match',
    guild=discord.Object(GUILD)
)
async def votemvp(interaction: discord.Interaction, player: discord.Member):
    global voting_in_progress, mvp_updates_today
    member = interaction.user

    try:
        async with aiosqlite.connect(DB_PATH) as conn:
            # Ensure the user has linked their Riot ID
            async with conn.execute("SELECT PlayerRiotID FROM PlayerStats WHERE DiscordID = ?", (str(member.id),)) as cursor:
                linked_account = await cursor.fetchone()
            if not linked_account or not linked_account[0]:
                await interaction.response.send_message("You must link your Riot ID before participating in MVP voting. Use `/link` to link your account.", ephemeral=True)
                return

            # Ensure voting is not exceeding the 3 MVP updates limit per day
            if mvp_updates_today >= 3:
                await interaction.response.send_message("The maximum number of MVP votes for today has been reached.", ephemeral=True)
                return

            # Ensure the user has the Player or Volunteer role
            if not any(role.name in ["Player", "Volunteer"] for role in member.roles):
                await interaction.response.send_message("You do not have the necessary role to vote.", ephemeral=True)
                return

            # Register the vote
            votes[player.display_name] += 1
            has_voted.add(member.id)

            # Notify everyone about the vote, mentioning both parties without notifications
            allowed_mentions = discord.AllowedMentions(users=False)
            await interaction.response.send_message(
                f"{member.mention} has voted for {player.mention}.",
                allowed_mentions=allowed_mentions,
                ephemeral=False
            )

            # If there's no voting session currently active, initiate one
            if not voting_in_progress:
                # Get Player and Volunteer roles
                player_role = discord.utils.get(interaction.guild.roles, name="Player")
                volunteer_role = discord.utils.get(interaction.guild.roles, name="Volunteer")
                if not player_role or not volunteer_role:
                    await interaction.followup.send("Player or Volunteer roles are not configured properly.", ephemeral=True)
                    return

                # Mention Player and Volunteer roles
                mentions = f"{player_role.mention} {volunteer_role.mention}"
                
                # Set allowed_mentions to explicitly mention roles
                allowed_mentions_roles = discord.AllowedMentions(roles=True)

                # Send voting initiation message with role mentions to everyone
                initial_message = await interaction.followup.send(
                    f"MVP voting has started! You have 5 minutes to vote using `/votemvp [username]`. {mentions}",
                    allowed_mentions=allowed_mentions_roles,
                    ephemeral=False
                )

                # Start the voting session loop
                start_voting.start(initial_message)

    except Exception as e:
        print(f"An error occurred: {e}")
        await interaction.response.send_message("An unexpected error occurred while processing your vote.", ephemeral=True)



# Admin command to cancel the voting session
@tree.command(
    name='cancelvoting',
    description='Cancel the current MVP voting session. (Admin Only)',
    guild=discord.Object(GUILD)
)
@commands.has_permissions(administrator=True)
async def cancelvoting(interaction: discord.Interaction):
    global voting_in_progress, votes, has_voted

    if voting_in_progress:
        start_voting.cancel()  # Cancel the ongoing voting loop
        voting_in_progress = False
        votes.clear()
        has_voted.clear()
        await interaction.response.send_message("The MVP voting session has been cancelled, and all votes have been discarded.", ephemeral=False)
    else:
        await interaction.response.send_message("There is no voting session currently active to cancel.", ephemeral=True)

# Admin command to finish the voting session immediately
@tree.command(
    name='finishvoting',
    description='Finish the current MVP voting session and award the MVP. (Admin Only)',
    guild=discord.Object(GUILD)
)
@commands.has_permissions(administrator=True)
async def finishvoting(interaction: discord.Interaction):
    if voting_in_progress:
        await finish_voting(interaction)
    else:
        await interaction.response.send_message("There is no voting session currently active to finish.", ephemeral=True)

# Function to finish voting and determine MVP
async def finish_voting(interaction):
    global voting_in_progress, mvp_updates_today, votes, has_voted

    try:
        async with aiosqlite.connect(DB_PATH) as conn:
            # Calculate MVP based on votes
            if votes:
                max_votes = max(votes.values())
                mvp_candidates = [player for player, count in votes.items() if count == max_votes]
                
                # Update MVPs in the database
                for mvp in mvp_candidates:
                    await conn.execute("UPDATE PlayerStats SET MVPs = MVPs + 1 WHERE DiscordUsername = ?", (mvp,))
                await conn.commit()

                # Prepare MVP result message
                if len(mvp_candidates) == 1:
                    await interaction.channel.send(f"🎉 {mvp_candidates[0]} has been voted the MVP of this round! 🎉")
                else:
                    mvp_list = ", ".join(mvp_candidates)
                    await interaction.channel.send(f"🎉 The MVP(s) with the highest votes are: {mvp_list}. 🎉")

                mvp_updates_today += 1
            else:
                await interaction.channel.send("No votes were cast. No MVP this round.")

    except Exception as e:
        print(f"An error occurred while finishing voting: {e}")
        await interaction.channel.send("An error occurred while processing the MVP votes.")

    finally:
        votes.clear()
        has_voted.clear()
        voting_in_progress = False


#end of mvp section

#help command section

# /help command with pagination
@tree.command(
    name='help',
    description='Get help for bot commands',
    guild=discord.Object(GUILD)
)
async def help_command(interaction: discord.Interaction):
    pages = [
        discord.Embed(
            title="Help Menu 📚",
            color=0xffc629
        ).add_field(
            name="/help",
            value="Displays documentation on all bot commands.",
            inline=False
        ).add_field(
            name="/link [riot_id]",
            value="Link your Riot ID to your Discord account. Users are required to type this command before any others except /help.",
            inline=False
        ).add_field(
            name="/rolepreference",
            value="Set your role preferences for matchmaking.",
            inline=False
        ),
        discord.Embed(
            title="Help Menu 📚",
            color=0xffc629
        ).add_field(
            name="/checkin",
            value="Initiate tournament check-in.",
            inline=False
        ).add_field(
            name="sitout",
            value="Volunteer to sit out of the current match..",
            inline=False
        ),
        discord.Embed(
            title="Help Menu 📚",
            description="**/wins [player_1] [player_2] [player_3] [player_4] [player_5]** - Add a win for the specified players.",
            color=0xffc629
        ),
        discord.Embed(
            title="Help Menu 📚",
            description="**/toxicity [discord_username]** - Give a user a point of toxicity.",
            color=0xffc629
        ),
        discord.Embed(
            title="Help Menu 📚",
            description="**/clear** - Remove all users from Player and Volunteer roles.",
            color=0xffc629
        ),
        discord.Embed(
            title="Help Menu 📚",
            description="**/players** - Find all players and volunteers currently enrolled in the game.",
            color=0xffc629
        ),
        discord.Embed(
            title="Help Menu 📚",
            description="**/points** - Update participation points in the spreadsheet.",
            color=0xffc629
        ),
        discord.Embed(
            title="Help Menu 📚",
            description="**/matchmake [match_number]** - Form teams for all players enrolled in the game.",
            color=0xffc629
        ),
        discord.Embed(
            title="Help Menu 📚",
            description="**/votemvp [username]** - Vote for the MVP of your match.",
            color=0xffc629
        ),
    ]

    # Set the footer for each page
    for i, page in enumerate(pages, start=1):
        page.set_footer(text=f"Page {i} of {len(pages)}")

    current_page = 0

    class HelpView(discord.ui.View):
        def __init__(self, *, timeout=180):
            super().__init__(timeout=timeout)
            self.message = None

        async def update_message(self, interaction: discord.Interaction):
            await interaction.response.defer()  # Defer response to prevent timeout error
            if self.message:
                await self.message.edit(embed=pages[current_page], view=self)

        @discord.ui.button(label="Previous", style=discord.ButtonStyle.primary)
        async def previous_page(self, interaction: discord.Interaction, button: discord.ui.Button):
            nonlocal current_page
            current_page = (current_page - 1) % len(pages)  # Loop back to last page if on the first page
            await self.update_message(interaction)

        @discord.ui.button(label="Next", style=discord.ButtonStyle.primary)
        async def next_page(self, interaction: discord.Interaction, button: discord.ui.Button):
            nonlocal current_page
            current_page = (current_page + 1) % len(pages)  # Loop back to first page if on the last page
            await self.update_message(interaction)

    view = HelpView()
    await interaction.response.send_message(embed=pages[current_page], view=view, ephemeral=True)
    view.message = await interaction.original_response()




        
# Shutdown of aiohttp session
async def close_session():
    global session
    if session is not None:
        await session.close()
        print("HTTP session has been closed.")

# Entry point to run async setup before bot starts
if __name__ == '__main__':
    try:
        # This line of code starts the bot.
        client.run(TOKEN)
    finally:
        asyncio.run(close_session())
        
        
        

"""
UNUSED / LEFTOVER CODE


The following code with the GatewayEventFilter class is a remnant of work done by a capstone team in spring 2024. We [the fall 2024] team made the decision to leave it
commented-out since that's how it was left for us, but it could possibly be useful in the future.

#Logger to catch discord disconects and ignores them.
class GatewayEventFilter(logging.Filter):
    def __init__(self) -> None:
        super().__init__('discord.gateway')
    def filter(self, record: logging.LogRecord) -> bool:
        if record.exc_info is not None and is instance(record.exc_info[1], discord.ConnectionClosed):
            return False
        return True

# The following commented-out line was left by the Spring 2024 capstone team, and though our fall 2024 version does not use it this may be useful for future groups.
#logging.getLogger('discord.gateway').addFilter(GatewayEventFilter())


The following is for a command, /viewvotes, that displayed an embed to users showing the number of votes for each MVP candidate. It was removed at the request of the sponsor
to prevent vote manipulation or a "bandwagon effect", and if it were re-added it may not work with the final version of MVP functionality submitted in the fall 2024 semester.

@tree.command(
    name='viewvotes',
    description='View the current MVP votes for the ongoing voting session.',
    guild=discord.Object(GUILD) 
)
async def viewvotes(interaction: discord.Interaction):
    global voting_in_progress
    if not voting_in_progress:
        await interaction.response.send_message("No voting session is currently active.", ephemeral=True)
        return

    # Create an embed to show the current votes
    embed = discord.Embed(title="🗳️ MVP Votes", color=0xffc629)

    # Add a field for each player who has received votes
    for player, count in votes.items():
        voters = [name for name in votes if votes[name] == count]
        voters_list = ', '.join(voters) if voters else "No votes yet"
        embed.add_field(name=player, value=f"Votes: {count}\n{voters_list}", inline=False)
    
    # Calculate remaining time in voting session
    remaining_time = start_voting.next_iteration - discord.utils.utcnow()
    minutes, seconds = divmod(int(remaining_time.total_seconds()), 60)
    embed.set_footer(text=f"Time remaining: {minutes} minutes and {seconds} seconds")

    await interaction.response.send_message(embed=embed, ephemeral=False)
"""