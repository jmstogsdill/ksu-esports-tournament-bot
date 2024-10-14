import discord
from discord.ext import commands
from openpyxl import load_workbook
import sqlite3

# Bot setup
intents = discord.Intents.default()
intents.message_content = True  # Enable message content intent
bot = commands.Bot(command_prefix="!", intents=intents)

# Path to Excel file and SQLite database
EXCEL_FILE_PATH = 'C:/Users/jknir/Downloads/capstone excel worksheet.xlsx'
DB_PATH = 'C:/Users/jknir/Downloads/ksu cap database/ksu_cap.db'

# SQLite connection function
def connect_db():
    conn = sqlite3.connect(DB_PATH)
    return conn

# Function to update the Excel file
def update_excel(player_name, wins, mvps, sheet_name):
    try:
        workbook = load_workbook(EXCEL_FILE_PATH)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f'Sheet {sheet_name} does not exist in the workbook')
        
        # Check if the player already exists in the sheet
        found = False
        for row in sheet.iter_rows(min_row=2):  # Assuming the first row is headers
            if row[0].value == player_name:
                row[1].value += wins  # Update wins
                row[2].value += mvps  # Update MVPs
                found = True
                break

        if not found:
            sheet.append([player_name, wins, mvps])  # Append new player if not found

        workbook.save(EXCEL_FILE_PATH)
    except Exception as e:
        print(f"Error updating Excel file: {e}")

# On bot ready event
@bot.event
async def on_ready():
    print(f'Bot is ready. Logged in as {bot.user}')

# Command to fetch player stats and update Excel

async def player_stats(ctx, player_name: str, sheet_name: str):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM Staging_PlayerStats WHERE Player=?", (player_name,))
    player = cursor.fetchone()

    if player:
        await ctx.send(f"Stats for {player_name}: Wins - {player[3]}, MVPs - {player[4]}")
        update_excel(player_name, player[3], player[4], sheet_name)
        await ctx.send(f"Player stats for {player_name} have been updated in the Excel sheet '{sheet_name}'.")
    else:
        await ctx.send(f"No stats found for {player_name}")

    conn.close()

# Command to record match result
@bot.command(name="record_match")
async def record_match(ctx, player_name: str, wins: int, mvps: int, sheet_name: str):
    conn = connect_db()
    cursor = conn.cursor()
    
    # Update the database with match results
    cursor.execute("UPDATE Staging_PlayerStats SET Wins = Wins + ?, MVPs = MVPs + ? WHERE Player = ?",
                   (wins, mvps, player_name))
    conn.commit()

    # Update the Excel file
    update_excel(player_name, wins, mvps, sheet_name)

    await ctx.send(f"Match results for {player_name} recorded. Wins: {wins}, MVPs: {mvps}.")
    conn.close()

# Run the bot with your Discord bot token
bot.run('MTI3ODk2MDQ0MzY1MzYyMzgxOA.Giu-dm.MWs00Unu751IJFJ9qJ-asre9iR2-bJnnhtNAj0')  # Replace with your actual Discord bot token
